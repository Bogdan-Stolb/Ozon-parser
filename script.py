import os
import re
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

class OzonParser:
    def __init__(self, headless=False, debug=True):
        self.debug = debug
        
        options = uc.ChromeOptions()
        
        if headless:
            options.add_argument('--headless=new')
        
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_argument('--disable-infobars')
        options.add_argument('--disable-extensions')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--window-size=1920,1080')
        options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36')
        
        options.add_experimental_option("prefs", {
            "credentials_enable_service": False,
            "profile.password_manager_enabled": False
        })
        
        self.driver = uc.Chrome(options=options, version_main=149)
        self.wait = WebDriverWait(self.driver, 15)
        
        if self.debug:
            print("[DEBUG] Браузер инициализирован")
            
    def _log(self, msg):
        if self.debug:
            print(f"[DEBUG] {msg}")
            
    def _clean_price(self, text):
        try:
            return int(re.sub(r'[^\d]', '', text))
        except:
            return None

    def search_and_parse(self, query, min_price, max_price, limit=None):
        self._log(f"Старт: '{query}', цена: {min_price}-{max_price}, лимит: {limit}")
        self.driver.get("https://www.ozon.ru/")
        time.sleep(3)
        
        try:
            search_input = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[name='text']")))
            search_input.click()
            time.sleep(1)
            search_input.send_keys(query)
            time.sleep(0.5)
            search_input.send_keys(Keys.ENTER)
            self._log("Запрос отправлен")
            time.sleep(4)
        except Exception as e:
            self._log(f"Ошибка поиска: {e}")
            return []

        data = []
        processed_urls = set()
        last_height = 0
        no_new_cards_count = 0
        search_url = self.driver.current_url
        
        while limit is None or len(data) < limit:
            # Собираем href И цену СРАЗУ, пока карточки свежие
            candidates = []
            try:
                cards = self.driver.find_elements(By.CSS_SELECTOR, "div.tile-root")
                self._log(f"Найдено карточек: {len(cards)}")
                
                for card in cards:
                    try:
                        # Получаем ссылку
                        link_el = card.find_element(By.CSS_SELECTOR, "a[href*='/product/']")
                        href = link_el.get_attribute('href')
                        if not href or href in processed_urls:
                            continue
                        
                        # Получаем цену - берём первый span с ценой внутри блока цены
                        price = None
                        try:
                            # Вариант 1: блок c35_4_0-a0 -> первый span
                            price_block = card.find_element(By.CSS_SELECTOR, "div[class*='c35_4_0-a0']")
                            price_spans = price_block.find_elements(By.CSS_SELECTOR, "span[class*='tsHeadline500Medium']")
                            if price_spans:
                                price = self._clean_price(price_spans[0].text)
                        except:
                            pass
                        
                        # Вариант 2: если не нашли - пробуем другой селектор
                        if not price:
                            try:
                                price_el = card.find_element(By.CSS_SELECTOR, "span[class*='tsHeadline500Medium']")
                                price = self._clean_price(price_el.text)
                            except:
                                pass
                        
                        if price:
                            self._log(f"  Карточка: {href.split('/')[-2][:30]}... | Цена: {price}₽")
                            candidates.append((href, price))
                        else:
                            self._log(f"  Карточка: цена не найдена для {href[:50]}")
                            
                    except Exception as e:
                        self._log(f"  Ошибка чтения карточки: {str(e)[:40]}")
                        continue
                        
            except Exception as e:
                self._log(f"Ошибка сбора карточек: {e}")
                break
            
            if not candidates:
                self._log("Кандидатов не найдено на этой странице")
            else:
                self._log(f"Кандидатов с ценой: {len(candidates)}")
            
            # Обрабатываем подходящие по цене
            for href, price in candidates:
                if limit and len(data) >= limit:
                    break
                
                if not (min_price <= price <= max_price):
                    self._log(f"  Пропуск: {price}₽ не в диапазоне {min_price}-{max_price}")
                    processed_urls.add(href)
                    continue
                
                processed_urls.add(href)
                self._log(f"✅ Подходящий товар: {price}₽ -> {href}")
                
                try:
                    # Переходим на страницу товара
                    self.driver.get(href)
                    time.sleep(2.5)
                    
                    item_data = self._parse_product_page(href, price)
                    if item_data:
                        data.append(item_data)
                        self._log(f"✅ Собрано: {len(data)}")
                    
                    # Возвращаемся к поиску
                    self.driver.get(search_url)
                    time.sleep(2.5)
                    
                except Exception as e:
                    self._log(f"Ошибка карточки: {str(e)[:50]}")
                    try:
                        self.driver.get(search_url)
                        time.sleep(2)
                    except:
                        pass
                    continue
            
            if limit and len(data) >= limit:
                break
                
            # Скролл для подгрузки новых карточек
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(3)
            
            new_height = self.driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                no_new_cards_count += 1
                if no_new_cards_count >= 3:
                    self._log("Достигнут конец выдачи")
                    break
            else:
                no_new_cards_count = 0
            last_height = new_height
            
        self._log(f"Итого собрано: {len(data)}")
        return data

    def _parse_product_page(self, url, card_price):
        item = {'URL': url, 'Цена_карточки': card_price, 'Название': '', 'Цена_страницы': '', 'Характеристики': {}}
        
        try:
            name_el = self.driver.find_element(By.CSS_SELECTOR, "div[data-widget='webProductHeading'] h1")
            item['Название'] = name_el.text.strip()
        except:
            pass
            
        try:
            price_el = self.driver.find_element(By.CSS_SELECTOR, "div[data-widget='webPrice'] span[class*='tsHeadline500Medium']")
            item['Цена_страницы'] = self._clean_price(price_el.text)
        except:
            try:
                price_el = self.driver.find_element(By.CSS_SELECTOR, "div[class*='pdp_b0i'] span")
                item['Цена_страницы'] = self._clean_price(price_el.text)
            except:
                pass

        try:
            char_blocks = self.driver.find_elements(By.CSS_SELECTOR, "div[data-widget='webShortCharacteristics'] div[class*='pdp_bo8']")
            for block in char_blocks:
                try:
                    key = block.find_element(By.XPATH, ".//span[contains(@style, 'textSecondary') or contains(@class, 'tsBodyM')]").text.strip()
                    val = block.find_element(By.XPATH, ".//span[contains(@style, 'textPrimary') or contains(@class, 'tsBody400Small')]").text.strip()
                    if key and val:
                        item['Характеристики'][key] = val
                except:
                    pass
        except:
            pass
            
        return item

    def export_to_excel(self, data, filename):
        if not data:
            self._log("Нет данных")
            return
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Товары"
        
        headers = ['№', 'Название', 'Цена (в выдаче)', 'Цена (на странице)', 'URL']
        all_keys = set()
        for item in data:
            all_keys.update(item['Характеристики'].keys())
        all_keys = sorted(list(all_keys))
        headers.extend(all_keys)
        
        ws.append(headers)
        hf = Font(bold=True, color="FFFFFF")
        hf_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for c in ws[1]:
            c.font, c.fill, c.alignment = hf, hf_fill, Alignment(horizontal="center")
            
        for i, item in enumerate(data, 1):
            row = [i, item['Название'], item['Цена_карточки'], item['Цена_страницы'], item['URL']]
            for key in all_keys:
                row.append(item['Характеристики'].get(key, ''))
            ws.append(row)
            
        for col in ws.columns:
            mx = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(mx + 2, 60)
            
        wb.save(filename)
        self._log(f"Excel сохранен: {filename}")

    def close(self):
        try:
            self.driver.quit()
        except:
            pass

def main():
    parser = OzonParser(headless=False, debug=True)
    try:
        q = input("Запрос: ").strip()
        mn = int(input("Мин. цена: ").strip())
        mx = int(input("Макс. цена: ").strip())
        lim = input("Лимит товаров (Enter = без лимита): ").strip()
        limit = int(lim) if lim else None
        
        data = parser.search_and_parse(q, mn, mx, limit)
        if data:
            fn = f"ozon_{q.replace(' ', '_')}_{mn}-{mx}.xlsx"
            parser.export_to_excel(data, fn)
            print(f"\nГотово: {fn}")
        else:
            print("Ничего не найдено")
    except KeyboardInterrupt:
        print("\nПрервано")
    except Exception as e:
        print(f"Ошибка: {e}")
    finally:
        parser.close()

if __name__ == "__main__":
    main()